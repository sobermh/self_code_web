from openpyxl.styles import Alignment, Font, numbers, PatternFill


class Autofit:
    def __init__(self, ws):
        """
        :param ws:  正在操作的表
        """
        self.ws = ws

    def autofit(self):
        from openpyxl import load_workbook, workbook
        from openpyxl.utils import get_column_letter

        # 第一步：计算每列最大宽度，并存储在列表lks中。

        lks = []  # 英文变量太费劲，用汉语首字拼音代替
        for i in range(1, self.ws.max_column + 1):  # 每列循环
            lk = 1  # 定义初始列宽，并在每个行循环完成后重置
            for j in range(1, self.ws.max_row + 1):  # 每行循环
                sz = self.ws.cell(row=j, column=i).value  # 每个单元格内容
                if isinstance(sz, str):  # 中文占用多个字节，需要分开处理
                    lk1 = len(sz.encode('utf-8'))  # gbk解码一个中文两字节，utf-8一个中文三字节，gbk合适
                else:
                    lk1 = len(str(sz))
                if lk < lk1:
                    lk = lk1  # 借助每行循环将最大值存入lk中
                # print(lk)
            lks.append(lk)  # 将每列最大宽度加入列表。（犯了一个错，用lks = lks.append(lk)报错，append会修改列表变量，返回值none，而none不能继续用append方法）

        # 第二步：设置列宽
        for i in range(1, self.ws.max_column + 1):
            k = get_column_letter(i)  # 将数字转化为列名,26个字母以内也可以用[chr(i).upper() for i in range(97, 123)]，不用导入模块
            self.ws.column_dimensions[k].width = lks[i - 1] + 2  # 设置列宽，一般加两个字节宽度，可以根据实际情况灵活调整

    @staticmethod
    def aligned_center_excel(ws):
        """
        单元格居重堆区
        :param ws:  正在操作的表
        :return:
        """
        center_aligned_text = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_aligned_text

    @staticmethod
    def bold_font_excel(ws, min_row, max_row):
        """
        加粗字体
        :param ws:  正在操作的表
        :param min_row:  起始行
        :param max_row:  结束行
        :return:
        """
        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=min_row, max_row=max_row):
            for cell in row:
                cell.font = bold_font

    @staticmethod
    def format_value_excel(ws, min_row, ex_col_list, decimal_format="0.00"):
        """
        格式化数值
        :param ws:  正在操作的表
        :param min_row:  起始行
        :param ex_col_list:  不需要格式化的列
        :param decimal_format:  格式化方式
        :return:
        """
        scientific_format = '0.00E+00'
        for row in ws.iter_rows(min_row=min_row):
            for cell in row:
                if cell.col_idx in ex_col_list:
                    continue
                try:
                    cell.value = float(cell.value)
                    if abs(cell.value) < 1e-2:
                        cell.number_format = scientific_format
                    else:
                        cell.number_format = decimal_format
                except Exception as e:
                    pass

    @staticmethod
    def text_to_number(ws, min_row, min_col):
        """
        文本转数值
        :param ws:  正在操作的表
        :param min_row:  起始行
        :param min_col:  起始列
        :return:
        """
        for row in ws.iter_rows(min_row=min_row, min_col=min_col):
            for cell in row:
                try:
                    cell.value = int(cell.value)
                except ValueError:
                    try:
                        cell.value = float(cell.value)
                    except ValueError:
                        pass
                except TypeError:
                    pass

    @staticmethod
    def fill_yellow_cell(ws, min_row, max_row):
        """
        填充黄色
        :param ws:  正在操作的表
        :param min_row:  起始行
        :param max_row:  结束行
        :return:
        """
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in ws.iter_rows(min_row=min_row, max_row=max_row):
            for cell in row:
                cell.font = yellow_fill
