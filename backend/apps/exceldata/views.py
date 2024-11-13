import re
import threading
import urllib.parse
from io import BytesIO
from itertools import islice, groupby
from operator import itemgetter

import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.worksheet.protection import SheetProtection
from rest_framework import status, permissions
from rest_framework.generics import GenericAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet

from apps.exceldata.filter import *
from apps.exceldata.serializers import *
from apps.exceldata.utils.autofit_excel import Autofit
from apps.exceldata.utils.download_execl import DownloadExcel


# Create your views here.
class CustomPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method == 'GET':
            return True
        return request.user and request.user.is_authenticated


class CustomPageNumberPagination(PageNumberPagination):
    def get_page_size(self, request):
        if 'page_size' in request.query_params:
            return int(request.query_params['page_size'])
            # return min(int(request.query_params['page_size']), self.max_page_size)
        return self.page_size

    # def get_paginated_response(self, data):
    #     return Response({
    #         'links': {
    #             'next': self.get_next_link(),
    #             'previous': self.get_previous_link()
    #         },
    #         'count': self.page.paginator.count,
    #         'page_size': self.get_page_size(self.request),
    #         'results': data
    #     })


# ---------------------- 重金属浓度2000-2030-----------------------
class MetalDensityView(ModelViewSet):
    queryset = MetalDensityModel.objects.all().order_by("id")  # 模型类queryset
    serializer_class = MetalDensitySerializer
    filterset_class = MetalDensityFilter
    pagination_class = CustomPageNumberPagination
    permission_classes = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):

        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)
            insert_data = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                    row_data_list = list(row_data.items())
                    for k, v in enumerate(row_data_list):
                        if k == 0:
                            continue

                        year = int("".join(re.findall(r'\d', row_data_list[0][1])))
                        city = "".join(re.findall(r'\D', row_data_list[0][1]))
                        insert_data.append(MetalDensityModel(year=year, city=city, type=v[0], value=v[1]))

            def batch_insert(type_model, iterable, batch_size):
                it = iter(iterable)
                while True:
                    chunk = list(islice(it, batch_size))
                    if not chunk:
                        return
                    type_model.objects.bulk_create(chunk)

            batch_insert(MetalDensityModel, insert_data, 1000)
            response_data = {
                "code": status.HTTP_201_CREATED,
                "message": "创建成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_201_CREATED)

        response_data = {
            "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
            "message": "创建失败",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        # print(response_del)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            MetalDensityModel.objects.all().delete()
            MetalCancerRiskModel.objects.all().delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            # 捕获并返回异常信息
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class MetalDensityDownloadView(GenericAPIView):
    queryset = MetalDensityModel.objects.all()  # 模型类queryset
    serializer_class = MetalDensitySerializer  # 序列化器

    def post(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "重金属浓度2000-2030"

        header = ["", "Cd", "Pd", "Cu", "Zn", "Ni", "Cr", "As", "Hg", "ΣHM"]
        insert_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('year', 'city'))
            for key, group in groups:
                year_city = f"{key[0]}{key[1]}"
                row_data = [year_city]
                for item in list(group):
                    row_data.append(item['value'])
                insert_data.append(row_data)

        process_data(self.queryset)
        df = pd.DataFrame(insert_data)
        ws.append(header)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        Autofit.aligned_center_excel(ws)
        Autofit.text_to_number(ws, 2, 2)
        Autofit.bold_font_excel(ws, 1, 1)
        Autofit(ws).autofit()

        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头

        response = DownloadExcel.gen_response(output, "重金属浓度2000-2030.xlsx")
        return response


# ---------------------- 重金属致癌2000-2030 -----------------------
class MetalCancerRiskView(ModelViewSet):
    queryset = MetalCancerRiskModel.objects.all().order_by("id")
    serializer_class = MetalCancerRiskSerializer
    filterset_class = MetalCancerRiskFilter
    pagination_class = CustomPageNumberPagination
    permission_classes = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        def batch_insert(type_model, iterable, batch_size):
            it = iter(iterable)
            while True:
                chunk = list(islice(it, batch_size))
                if not chunk:
                    return
                type_model.objects.bulk_create(chunk)

        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)

            insert_data = []
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if sheet_name in ["长三角重金属儿童致癌", "长三角重金属成人致癌"]:
                    for index, row in df.iterrows():
                        row_data = row.to_dict()
                        row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                        if index == 0:
                            continue
                        for key in row_data.keys():
                            if key in ["Unnamed: 0", "Unnamed: 1"]:
                                continue
                            city = row_data.get("Unnamed: 0")
                            year = row_data.get("Unnamed: 1")
                            value_str = row_data.get(key, "")
                            value_str = re.sub(r'\s+', '', str(value_str))  # Remove all whitespace characters.
                            if sheet_name == "长三角重金属儿童致癌":
                                insert_data.append(
                                    MetalCancerRiskModel(year=year, city=city, type=key, value=value_str,
                                                         is_adult=False))
                            else:
                                insert_data.append(
                                    MetalCancerRiskModel(year=year, city=city, type=key, value=value_str,
                                                         is_adult=True))

            batch_insert(MetalCancerRiskModel, insert_data, 1000)
            response_data = {
                "code": status.HTTP_201_CREATED,
                "message": "创建成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_201_CREATED)
        response_data = {
            "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
            "message": "创建失败",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        # print(response_del)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            MetalCancerRiskModel.objects.all().delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class MetalCancerRiskDownloadView(GenericAPIView):
    queryset = MetalCancerRiskModel.objects.all().order_by("id")
    serializer_class = MetalCancerRiskSerializer

    def post(self, request):
        wb = Workbook()
        ws_child = wb.active
        ws_child.title = "长三角重金属儿童致癌"
        ws_adult = wb.create_sheet("长三角重金属成人致癌")
        ws_param = wb.create_sheet("暴露参数表")

        header = ["", "", "经口摄入", "皮肤接触", "呼吸吸入", "总摄入"]
        type_data = ["地区", "时间", "TCRing", "TCRder", "Crinh", "TCR"]
        adult_data = []
        child_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('city', 'year', 'is_adult'))
            for key, group in groups:
                row_data = [key[0], key[1]]
                for item in list(group):
                    row_data.append(item['value'])
                if key[2]:
                    adult_data.append(row_data)
                else:
                    child_data.append(row_data)

        process_data(self.queryset)
        child_df = pd.DataFrame(child_data)
        adult_df = pd.DataFrame(adult_data)
        for ws in [ws_child, ws_adult]:
            ws.append(header)
            ws.append(type_data)
            df = child_df if ws == ws_child else adult_df
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)
            Autofit.aligned_center_excel(ws)
            Autofit.format_value_excel(ws, 2, [1, 2])
            Autofit.bold_font_excel(ws, 1, 2)
            Autofit(ws).autofit()

        ws_param.append(["", "", "", "参考值", "", "", "", "", "", "参考值"])
        ws_param.append(["", "", "参数", "儿童", "成人", "", "", "参数", "单位", "儿童", "成人"])
        ws_param.append(
            ["", "", "TCRing", 200, 100, "", "", "土壤颗粒摄入速率 Ingestion rate（IRing）", "mg/d", 200, 100])
        ws_param.append(
            ["", "", "IRinh", 7.5, 14.5, "", "", "土壤颗粒吸入速率 Inhalation rate（IRinh）", f"m\u00B3/d", 7.5, 14.5])
        ws_param.append(["", "", "EF", 350, 350, "", "", "暴露频率 Exposure frequency（EF）", "day/a", 350, 350])
        ws_param.append(["", "", "ED", 6, 24, "", "", "暴露期 Exposure duration（ED）", "a", 6, 24])
        ws_param.append(
            ["", "", "SA", 2448, 5075, "", "", "皮肤暴露面积 Exposed skin area（SA）", f"cm\u00B2", 2448, 5075])
        ws_param.append(
            ["", "", "AF", 0.2, 0.07, "", "", "皮肤黏附系数 Skin adherence factor（AF）]", f"mg*m\u207B\u00B2*d", 0.2,
             0.07])
        ws_param.append(
            ["", "", "ABS", 0.001, 0.001, "", "", "皮肤吸收因子 Dermal absorption factor（ ABS）", "", 0.001, 0.001])
        ws_param.append(["", "", "PEF", 1.36E+09, 1.36E+09, "", "", "颗粒物释放因子 Particle emission factor（PEF）",
                         "m\u00B3*kg\u207B\u00B9", 1.36E+09, 1.36E+09])
        ws_param.append(
            ["", "", "AT", 25550, 25550, "", "", "平均暴露时间 Average exposure time（AT）", "d", 25550, 25550])
        ws_param.append(["", "", "", 2190, 8760, "", "", "", "", 2190, 8760])
        ws_param.append(["", "", "BW", 15.9, 56.8, "", "", "平均体重 Average bodyweight（BW）", "kg", 15.9, 56.8])
        ws_param.append([])
        ws_param.append([])
        ws_param.append(["", "", "", "儿童", "成人"])
        ws_param.append(["", "ADDing", "致癌", 1.03385887826312E-06, 5.7881535790083E-07])
        ws_param.append(["", "", "非致癌", 0.0000120616869130697, 1.68821146054409E-06])
        ws_param.append(["", "ADDinh", "致癌", 2.85071381874021E-11, 6.17119315408973E-11])
        ws_param.append(["", "", "非致癌", 3.32583278853025E-10, 1.7999313366095E-10])
        ws_param.append(["", "ADDder", "致癌", 2.53088653398811E-09, 2.0562415589427E-09])
        ws_param.append(["", "", "非致癌", 2.95270095631946E-08, 5.99737121358287E-09])
        ws_param.append(["", "ADDplant", "致癌"])
        ws_param.append(["", "", "非致癌"])
        ws_param.append([])
        ws_param.append([])
        ws_param.append(["", "", "Rfd"])
        ws_param.append(["", "元素", "经口摄入", "皮肤接触", "呼吸吸入"])
        ws_param.append(["", "", "ing", "der", "inh"])
        ws_param.append(["", "Cd", 0.001, 0.00001, 0.00001])
        ws_param.append(["", "Cr", 0.003, 0.00006, 0.0000286])
        ws_param.append(["", "As", 0.0003, 0.000123])
        ws_param.append(["", "Pb", 0.0035, 0.000525])
        ws_param.append(["", "Cu", 0.04, 0.012])
        ws_param.append(["", "Zn", 0.3, 0.06])
        ws_param.append(["", "Ni", 0.02, 0.0054, 0.00009])
        ws_param.append(["", "Hg", 0.0003, 0.000021, 0.0000857])
        ws_param.append([])
        ws_param.append([])
        ws_param.append(["", "", "SF"])
        ws_param.append(["", "元素", "经口摄入", "皮肤接触", "呼吸吸入"])
        ws_param.append(["", "Cd", 6.1, "", 6.3])
        ws_param.append(["", "Cr", 0.5, "", 42])
        ws_param.append(["", "As", 1.5, 3.66, 15])
        ws_param.append(["", "Pb", 0.0085])
        ws_param.append(["", "Cu"])
        ws_param.append(["", "Zn"])
        ws_param.append(["", "Ni", "", "", 0.84])
        ws_param.append(["", "Hg"])
        Autofit.aligned_center_excel(ws_param)
        Autofit(ws_param).autofit()
        protection = SheetProtection()
        protection.enable()
        ws_param.protection = protection

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头
        response = DownloadExcel.gen_response(output, "重金属致癌2000-2030.xlsx")
        return response


# ---------------------- 重金属非致癌2000-2030 -----------------------
class MetalNonCancerRiskView(ModelViewSet):
    queryset = MetalNonCancerRiskModel.objects.all().order_by("id")
    serializer_class = MetalNonCancerRiskSerializer
    filterset_class = MetalNonCancerRiskFilter
    pagination_class = CustomPageNumberPagination
    permission_classes = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        def batch_insert(type_model, iterable, batch_size):
            it = iter(iterable)
            while True:
                chunk = list(islice(it, batch_size))
                if not chunk:
                    return
                type_model.objects.bulk_create(chunk)

        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)

            insert_data = []
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if sheet_name in ["长三角重金属儿童非致癌", "长三角重金属成人非致癌"]:
                    for index, row in df.iterrows():
                        row_data = row.to_dict()
                        row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                        if index == 0:
                            continue
                        for key in row_data.keys():
                            if key in ["Unnamed: 0", "Unnamed: 1"]:
                                continue
                            city = row_data.get("Unnamed: 0")
                            year = row_data.get("Unnamed: 1")
                            value_str = row_data.get(key, "")
                            value_str = re.sub(r'\s+', '', str(value_str))  # Remove all whitespace characters.
                            if sheet_name == "长三角重金属儿童非致癌":
                                insert_data.append(
                                    MetalNonCancerRiskModel(year=year, city=city, type=key, value=value_str,
                                                            is_adult=False))
                            else:
                                insert_data.append(
                                    MetalNonCancerRiskModel(year=year, city=city, type=key, value=value_str,
                                                            is_adult=True))

            batch_insert(MetalNonCancerRiskModel, insert_data, 1000)
            response_data = {
                "code": status.HTTP_201_CREATED,
                "message": "创建成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_201_CREATED)
        response_data = {
            "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
            "message": "创建失败",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        # print(response_del)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            MetalNonCancerRiskModel.objects.all().delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class MetalNonCancerRiskDownloadView(GenericAPIView):
    queryset = MetalNonCancerRiskModel.objects.all().order_by("id")
    serializer_class = MetalNonCancerRiskSerializer

    def post(self, request):
        wb = Workbook()
        ws_child = wb.active
        ws_child.title = "长三角重金属儿童非致癌"
        ws_adult = wb.create_sheet("长三角重金属成人非致癌")
        ws_param = wb.create_sheet("暴露参数表")

        header = ["", "", "经口摄入", "皮肤接触", "呼吸吸入", "总摄入"]
        type_data = ["地区", "时间", "Hiing", "Hider", "Hiinh", "THI"]
        adult_data = []
        child_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('city', 'year', 'is_adult'))
            for key, group in groups:
                row_data = [key[0], key[1]]
                for item in list(group):
                    row_data.append(item['value'])
                if key[2]:
                    adult_data.append(row_data)
                else:
                    child_data.append(row_data)

        process_data(self.queryset)
        child_df = pd.DataFrame(child_data)
        adult_df = pd.DataFrame(adult_data)
        for ws in [ws_child, ws_adult]:
            ws.append(header)
            ws.append(type_data)
            df = child_df if ws == ws_child else adult_df
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)
            Autofit.aligned_center_excel(ws)
            Autofit.format_value_excel(ws, 2, [1, 2])
            Autofit.bold_font_excel(ws, 1, 2)
            Autofit(ws).autofit()

        ws_param.append(["", "", "", "参考值"])
        ws_param.append(["参数", "单位", "儿童", "成人"])
        ws_param.append(["土壤颗粒摄入速率 Ingestion rate（IRing）", "mg/d", 200, 100])
        ws_param.append(["土壤颗粒吸入速率 Inhalation rate（IRinh）", f"m\u00B3/d", 7.5, 14.5])
        ws_param.append(["暴露频率 Exposure frequency（EF）", "day/a", 350, 350])
        ws_param.append(["暴露期 Exposure duration（ED）", "a", 6, 24])
        ws_param.append(["皮肤暴露面积 Exposed skin area（SA）", f"cm\u00B2", 2448, 5075])
        ws_param.append(["皮肤黏附系数 Skin adherence factor（AF）]", f"mg*m\u207B\u00B2*d", 0.2, 0.07])
        ws_param.append(["皮肤吸收因子 Dermal absorption factor（ ABS）", "", 0.001, 0.001])
        ws_param.append(["颗粒物释放因子 Particle emission factor（PEF）", "m\u00B3*kg\u207B\u00B9", 1.36E+09, 1.36E+09])
        ws_param.append(["平均暴露时间 Average exposure time（AT）", "d", 25550, 25550])
        ws_param.append(["", "", 2190, 8760])
        ws_param.append(["平均体重 Average bodyweight（BW）", "kg", 15.9, 56.8])
        Autofit.aligned_center_excel(ws_param)
        Autofit(ws_param).autofit()
        protection = SheetProtection()
        protection.enable()
        ws_param.protection = protection

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头
        response = DownloadExcel.gen_response(output, "重金属非致癌2000-2030.xlsx")
        return response


# ---------------------- 重金属每年平均数据-----------------------
class MetalAverageAnnualView(ModelViewSet):
    queryset = MetalAverageAnnualModel.objects.all().order_by("id")
    serializer_class = MetalAverageAnnualSerializer
    filterset_class = MetalAverageAnnualFilter
    pagination_class = CustomPageNumberPagination
    permissions = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)
            insert_data = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                    row_data_list = list(row_data.items())
                    for k, v in enumerate(row_data_list):
                        if k == 0:
                            continue
                        year = int(row_data_list[0][1])
                        insert_data.append(MetalAverageAnnualModel(year=year, type=v[0], value=v[1]))

            def batch_insert(type_model, iterable, batch_size):
                it = iter(iterable)
                while True:
                    chunk = list(islice(it, batch_size))
                    if not chunk:
                        return
                    type_model.objects.bulk_create(chunk)

            batch_insert(MetalAverageAnnualModel, insert_data, 1000)

        response_data = {
            "code": status.HTTP_201_CREATED,
            "message": "创建成功",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            self.queryset.delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            # 捕获并返回异常信息
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class MetalAverageAnnualDownloadView(GenericAPIView):
    queryset = MetalAverageAnnualModel.objects.all().order_by("id")
    serializer_class = MetalAverageAnnualSerializer

    def post(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "重金属每年平均数据"

        header = ["year", "Cd", "As", "Pb", "Zn", "Hg", "Cu", "Ni", "Cr"]
        insert_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('year'))
            for key, group in groups:
                group_data = list(group)
                row_data = [key]
                for item in group_data:
                    row_data.append(item['value'])
                insert_data.append(row_data)

        process_data(self.queryset)
        df = pd.DataFrame(insert_data)
        ws.append(header)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        Autofit.aligned_center_excel(ws)
        Autofit.format_value_excel(ws, 2, [1])
        Autofit.bold_font_excel(ws, 1, 1)
        Autofit(ws).autofit()

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头

        response = DownloadExcel.gen_response(output, "重金属每年平均数据.xlsx")
        return response


# ---------------------- 重金属数据量-----------------------
class MetalReferenceView(ModelViewSet):
    queryset = MetalReferenceModel.objects.all().order_by("id")
    serializer_class = MetalReferenceSerializer
    filterset_class = MetalReferenceFilter
    pagination_class = CustomPageNumberPagination
    permissions = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)
            insert_data = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                    row_data_list = list(row_data.items())
                    for k, v in enumerate(row_data_list):
                        if k == 0:
                            continue
                        value = int(row_data_list[1][1])
                        insert_data.append(MetalReferenceModel(type=row_data_list[0][1], reference_number=value))

            def batch_insert(type_model, iterable, batch_size):
                it = iter(iterable)
                while True:
                    chunk = list(islice(it, batch_size))
                    if not chunk:
                        return
                    type_model.objects.bulk_create(chunk)

            batch_insert(MetalReferenceModel, insert_data, 1000)

        response_data = {
            "code": status.HTTP_201_CREATED,
            "message": "创建成功",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            self.queryset.delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            # 捕获并返回异常信息
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class MetalReferenceDownloadView(GenericAPIView):
    queryset = MetalReferenceModel.objects.all().order_by("id")
    serializer_class = MetalReferenceSerializer

    def post(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "重金属数据量"

        header = ["种类", "文献数量"]
        insert_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('type'))
            for key, group in groups:
                group_data = list(group)
                row_data = [key]
                for item in group_data:
                    row_data.append(item['reference_number'])
                insert_data.append(row_data)

        process_data(self.queryset)
        df = pd.DataFrame(insert_data)
        ws.append(header)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        Autofit.aligned_center_excel(ws)
        Autofit.bold_font_excel(ws, 1, 1)
        Autofit(ws).autofit()

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头

        response = DownloadExcel.gen_response(output, "重金属数据量.xlsx")
        return response


# ---------------------- 社会因子数据2000-2030 -----------------------
class SocialEconomyView(ModelViewSet):
    queryset = SocialEconomyModel.objects.all().order_by("id")  # 模型类queryset
    serializer_class = SocialEconomySerializer
    filterset_class = SocialEconomyFilter
    pagination_class = CustomPageNumberPagination
    permission_classes = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        def batch_insert(iterable, batch_size):
            it = iter(iterable)
            while True:
                chunk = list(islice(it, batch_size))
                if not chunk:
                    return
                SocialEconomyModel.objects.bulk_create(chunk)

        file = request.FILES.get('file')
        if file:
            # 使用 pandas 读取 Excel 文件
            xls = pd.ExcelFile(file)

            insert_data = []
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                type_dict = {}
                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                    # print(row_data)
                    if index == 0:
                        type_dict = row_data
                        continue
                    # year_str = row_data.get('Unnamed: 1')
                    # year = int(year_str.replace("年", ""))
                    # city = row_data.get('Unnamed: 0')
                    year = ""
                    factors = ""
                    city = ""
                    for key in row_data.keys():
                        if key in ["Unnamed: 0", "Unnamed: 1"]:
                            continue
                        if "经济" in key:
                            factors = "经济"
                        if "农业" in key:
                            factors = "农业"
                        if "工业" in key:
                            factors = "工业"
                        if "交通" in key:
                            factors = "交通"
                        city = row_data.get("Unnamed: 0")
                        year = row_data.get("Unnamed: 1")
                        type1 = type_dict.get(key)
                        value_str = row_data.get(key, "")
                        value_str = re.sub(r'\s+', '', str(value_str))  # Remove all whitespace characters.
                        insert_data.append(
                            SocialEconomyModel(year=year, city=city, type=type1, value=value_str, factors=factors))
            batch_insert(insert_data, 1000)
            response_data = {
                "code": status.HTTP_201_CREATED,
                "message": "创建成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_201_CREATED)
        response_data = {
            "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
            "message": "创建失败",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        # print(response_del)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            SocialEconomyModel.objects.all().delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class SocialEconomyDownloadView(GenericAPIView):
    queryset = SocialEconomyModel.objects.all()  # 模型类queryset
    serializer_class = SocialEconomySerializer  # 序列化器

    def post(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "社会因子数据2000-2030"

        factors_data = ["", "", "经济", "经济", "经济", "经济", "经济", "经济", "农业", "农业", "农业", "农业", "农业",
                        "农业", "工业", "工业", "工业", "工业", "交通", "交通", "交通", "交通"]
        type_data = ["地区", "时间", "第一产业总产值(亿元)", "第二产业总产值(亿元)", "第三产业总产值(亿元)",
                     "种植业产值（亿元）", "林业产值（亿元）", "牧业产值（亿元）", "农用化肥施用折纯量(吨)",
                     "农用氮肥施用折纯量(吨)", "农用磷肥施用折纯量(吨)",
                     "农用钾肥施用折纯量(吨)", "农用复合肥施用折纯量(吨)", "农药使用量(吨)",
                     "建筑业总产值（亿元）", "规模以上工业总产值(亿元)", "轻工业（亿元）", "重工业（亿元）",
                     "货运量(万吨)", "铁路货运量(万吨)", "公路货运量(万吨)", "水运货运量(万吨)",
                     ]
        insert_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('city', 'year'))
            for key, group in groups:
                row_data = [key[0], f"{key[1]}"]
                for item in list(group):
                    row_data.append(item['value'])
                insert_data.append(row_data)

        process_data(self.queryset)

        df = pd.DataFrame(insert_data)
        ws.append(factors_data)
        ws.append(type_data)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        Autofit.aligned_center_excel(ws)
        Autofit.text_to_number(ws, 3, 2)
        Autofit.bold_font_excel(ws, 1, 2)
        Autofit(ws).autofit()

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头

        response = DownloadExcel.gen_response(output, "社会因子数据2000-2030.xlsx")
        return response


# --------------------- 多环芳烃每年平均值 ---------------------

class PAHsAverageAnnualView(ModelViewSet):
    queryset = PAHsAverageAnnualModel.objects.all().order_by("id")  # 模型类queryset
    serializer_class = PAHsAverageAnnualSerializer
    filterset_class = PAHsAverageAnnualFilter
    pagination_class = CustomPageNumberPagination
    permission_classes = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)
            insert_data = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                    row_data_list = list(row_data.items())
                    for k, v in enumerate(row_data_list):
                        if k == 0:
                            continue
                        year = int(row_data_list[0][1])
                        insert_data.append(PAHsAverageAnnualModel(year=year, type=v[0], value=v[1]))

            def batch_insert(type_model, iterable, batch_size):
                it = iter(iterable)
                while True:
                    chunk = list(islice(it, batch_size))
                    if not chunk:
                        return
                    type_model.objects.bulk_create(chunk)

            batch_insert(PAHsAverageAnnualModel, insert_data, 1000)

        response_data = {
            "code": status.HTTP_201_CREATED,
            "message": "创建成功",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            self.queryset.delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            # 捕获并返回异常信息
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class PAHsAverageAnnualDownloadView(GenericAPIView):
    queryset = PAHsAverageAnnualModel.objects.all().order_by("id")
    serializer_class = PAHsAverageAnnualSerializer

    def post(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "多环芳烃每年平均值"

        header = ["year", "LMW PAHs", "MMW PAHs", "HMW PAHs", "Σ16PAHs"]
        insert_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('year'))
            for key, group in groups:
                group_data = list(group)
                row_data = [key]
                for item in group_data:
                    row_data.append(item['value'])
                insert_data.append(row_data)

        process_data(self.queryset)
        df = pd.DataFrame(insert_data)
        ws.append(header)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        Autofit.aligned_center_excel(ws)
        Autofit.format_value_excel(ws, 2, [1])
        Autofit.bold_font_excel(ws, 1, 1)
        Autofit(ws).autofit()

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头

        response = DownloadExcel.gen_response(output, "多环芳烃每年平均值.xlsx")
        return response


# --------------------- 土地利用率占比 ---------------------
class LandUseRatioView(ModelViewSet):
    queryset = LandUseRatioModel.objects.all().order_by("id")
    serializer_class = LandUseRatioSerializer
    filterset_class = LandUseRatioFilter
    pagination_class = CustomPageNumberPagination
    permissions = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)
            insert_data = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                    row_data_list = list(row_data.items())
                    for k, v in enumerate(row_data_list):
                        # print(k,v)
                        if k == 0:
                            continue
                        type1 = row_data_list[0][1]
                        insert_data.append(LandUseRatioModel(type=type1, value=v[1]))

            def batch_insert(type_model, iterable, batch_size):
                it = iter(iterable)
                while True:
                    chunk = list(islice(it, batch_size))
                    if not chunk:
                        return
                    type_model.objects.bulk_create(chunk)

            batch_insert(LandUseRatioModel, insert_data, 1000)

        response_data = {
            "code": status.HTTP_201_CREATED,
            "message": "创建成功",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            self.queryset.delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            # 捕获并返回异常信息
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class LandUseRatioDownloadView(GenericAPIView):
    queryset = LandUseRatioModel.objects.all().order_by("id")
    serializer_class = LandUseRatioSerializer

    def post(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "土地利用面积占比"

        header = ["土地类型", "面积(平方千米)"]
        insert_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('type'))
            for key, group in groups:
                group_data = list(group)
                row_data = [key]
                for item in group_data:
                    row_data.append(item['value'])
                insert_data.append(row_data)

        process_data(self.queryset)
        df = pd.DataFrame(insert_data)
        ws.append(header)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        Autofit.aligned_center_excel(ws)
        Autofit.format_value_excel(ws, 2, [1])
        Autofit.bold_font_excel(ws, 1, 1)
        Autofit(ws).autofit()

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头

        response = DownloadExcel.gen_response(output, "土地利用面积占比.xlsx")
        return response


# --------------------- 多环芳烃数据量 ---------------------
class PAHsReferenceView(ModelViewSet):
    queryset = PAHsReferenceModel.objects.all().order_by("id")
    serializer_class = PAHsReferenceSerializer
    filterset_class = PAHsReferenceFilter
    pagination_class = CustomPageNumberPagination
    permissions = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)
            insert_data = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                    row_data_list = list(row_data.items())
                    for k, v in enumerate(row_data_list):
                        if k == 0:
                            continue
                        value = int(row_data_list[1][1])
                        insert_data.append(PAHsReferenceModel(year=row_data_list[0][1], reference_number=value))

            def batch_insert(type_model, iterable, batch_size):
                it = iter(iterable)
                while True:
                    chunk = list(islice(it, batch_size))
                    if not chunk:
                        return
                    type_model.objects.bulk_create(chunk)

            batch_insert(PAHsReferenceModel, insert_data, 1000)

        response_data = {
            "code": status.HTTP_201_CREATED,
            "message": "创建成功",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            self.queryset.delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            # 捕获并返回异常信息
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class PAHsReferenceDownloadView(GenericAPIView):
    queryset = PAHsReferenceModel.objects.all().order_by("id")
    serializer_class = PAHsReferenceSerializer

    def post(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "多环芳烃数据量"

        header = ["年份", "文献数量"]
        insert_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('year'))
            for key, group in groups:
                group_data = list(group)
                row_data = [key]
                for item in group_data:
                    row_data.append(item['reference_number'])
                insert_data.append(row_data)

        process_data(self.queryset)
        df = pd.DataFrame(insert_data)
        ws.append(header)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        Autofit.aligned_center_excel(ws)
        Autofit.bold_font_excel(ws, 1, 1)
        Autofit(ws).autofit()

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头

        response = DownloadExcel.gen_response(output, "多环芳烃数据量.xlsx")
        return response


# --------------------- 多环芳烃致癌风险2000-2030 ---------------------
class PAHsCancerRiskView(ModelViewSet):
    queryset = PAHsCancerRiskModel.objects.all().order_by("id")
    serializer_class = PAHsCancerRiskSerializer
    filterset_class = PAHsCancerRiskFilter
    pagination_class = CustomPageNumberPagination
    permission_classes = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        def batch_insert(type_model, iterable, batch_size):
            it = iter(iterable)
            while True:
                chunk = list(islice(it, batch_size))
                if not chunk:
                    return
                type_model.objects.bulk_create(chunk)

        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)

            insert_data = []
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if sheet_name in ["长三角多环芳烃致癌成人", "长三角多环芳烃致癌儿童"]:
                    for index, row in df.iterrows():
                        row_data = row.to_dict()
                        row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                        if index == 0:
                            continue
                        for key in row_data.keys():
                            if key in ["Unnamed: 0", "Unnamed: 1"]:
                                continue
                            city = row_data.get("Unnamed: 0")
                            year = row_data.get("Unnamed: 1")
                            value_str = row_data.get(key, "")
                            value_str = re.sub(r'\s+', '', str(value_str))  # Remove all whitespace characters.
                            if sheet_name == "长三角多环芳烃致癌儿童":
                                insert_data.append(
                                    PAHsCancerRiskModel(year=year, city=city, type=key, value=value_str,
                                                        is_adult=False))
                            else:
                                insert_data.append(
                                    PAHsCancerRiskModel(year=year, city=city, type=key, value=value_str,
                                                        is_adult=True))

            batch_insert(PAHsCancerRiskModel, insert_data, 1000)
            response_data = {
                "code": status.HTTP_201_CREATED,
                "message": "创建成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_201_CREATED)
        response_data = {
            "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
            "message": "创建失败",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        # print(response_del)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            self.get_queryset().delete()

            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class PAHsCancerRiskDownloadView(GenericAPIView):
    queryset = PAHsCancerRiskModel.objects.all().order_by("id")
    serializer_class = PAHsCancerRiskSerializer

    def post(self, request):
        wb = Workbook()
        ws_adult = wb.active
        ws_adult.title = "长三角多环芳烃致癌成人"
        ws_child = wb.create_sheet("长三角多环芳烃致癌儿童")

        header = ["", "", "经口摄入", "皮肤接触", "呼吸吸入", "总摄入"]
        type_data = ["地区", "时间", "ILCRing", "ILCRinh", "ILCRder", "TILCR"]
        adult_data = []
        child_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('city', 'year', 'is_adult'))
            for key, group in groups:
                row_data = [key[0], key[1]]
                for item in list(group):
                    row_data.append(item['value'])
                if key[2]:
                    adult_data.append(row_data)
                else:
                    child_data.append(row_data)

        process_data(self.queryset)
        child_df = pd.DataFrame(child_data)
        adult_df = pd.DataFrame(adult_data)
        for ws in [ws_child, ws_adult]:
            ws.append(header)
            ws.append(type_data)
            df = child_df if ws == ws_child else adult_df
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)
            Autofit.aligned_center_excel(ws)
            Autofit.format_value_excel(ws, 2, [1, 2])
            Autofit.bold_font_excel(ws, 1, 2)
            Autofit(ws).autofit()

        # 使用 BytesIO 来保存工作簿
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头
        response = DownloadExcel.gen_response(output, "多环芳烃致癌风险2000-2030.xlsx")
        return response


class PAHsDensityView(ModelViewSet):
    queryset = PAHsDensityModel.objects.all().order_by("id")  # 模型类queryset
    serializer_class = PAHsDensitySerializer
    filterset_class = PAHsDensityFilter
    pagination_class = CustomPageNumberPagination
    permission_classes = [CustomPermission]

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if file:
            xls = pd.ExcelFile(file)
            insert_data = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                for index, row in df.iterrows():
                    row_data = row.to_dict()
                    row_data = {k: (None if pd.isna(v) else v) for k, v in row_data.items()}
                    row_data_list = list(row_data.items())
                    for k, v in enumerate(row_data_list):
                        if k == 0:
                            continue
                        year = int("".join(re.findall(r'\d', row_data_list[0][1])))
                        city = "".join(re.findall(r'\D', row_data_list[0][1]))
                        insert_data.append(PAHsDensityModel(year=year, city=city, type=v[0], value=v[1]))

            def batch_insert(type_model, iterable, batch_size):
                it = iter(iterable)
                while True:
                    chunk = list(islice(it, batch_size))
                    if not chunk:
                        return
                    type_model.objects.bulk_create(chunk)

            batch_insert(PAHsDensityModel, insert_data, 1000)

            response_data = {
                "code": status.HTTP_201_CREATED,
                "message": "创建成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_201_CREATED)

        response_data = {
            "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
            "message": "创建失败",
            "data": ""
        }
        return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        response_del = self.destroy(request, *args, **kwargs)
        if response_del.status_code == 200:
            return self.create(request, *args, **kwargs)
        return response_del

    def destroy(self, request, *args, **kwargs):
        try:
            # 删除所有记录
            self.get_queryset().delete()
            response_data = {
                "code": status.HTTP_200_OK,
                "message": "删除成功",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            response_data = {
                "code": status.HTTP_500_INTERNAL_SERVER_ERROR,
                "message": f"删除失败: {str(e)}",
                "data": ""
            }
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        all_param = str(request.query_params.get('all', 0)) == "1"
        queryset = self.filter_queryset(self.get_queryset())
        if all_param:
            self.paginator.page_size = queryset.count()
        return super().list(request, *args, **kwargs)


class PAHsDensityDownloadView(GenericAPIView):
    queryset = PAHsDensityModel.objects.all()
    serializer_class = PAHsDensitySerializer

    def post(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "多环芳烃浓度2000-2030"

        header = ["", "Nap(μg/kg)", "Acy(μg/kg)", "Ace(μg/kg)", "Flu(μg/kg)", "Phe(μg/kg)",
                  "Ant(μg/kg)", "Fla(μg/kg)", "Pyr(μg/kg)", "BaA(μg/kg)", "Chr(μg/kg)", "BbF(μg/kg)",
                  "BkF(μg/kg)",
                  "BaP(μg/kg)", "DahA(μg/kg)", "BghiP(μg/kg)", "InP(μg/kg)", "Σ16PAHs"]
        insert_data = []

        def process_data(data):
            groups = groupby(data.values(), key=itemgetter('year', 'city'))
            for key, group in groups:
                year_city = f"{key[0]}{key[1]}"
                row_data = [year_city]
                for item in list(group):
                    row_data.append(item['value'])
                insert_data.append(row_data)

        process_data(self.queryset)
        df = pd.DataFrame(insert_data)
        ws.append(header)
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        Autofit.aligned_center_excel(ws)
        Autofit.text_to_number(ws, 2, 2)
        Autofit.format_value_excel(ws, 2, [1], "0.000")
        Autofit.bold_font_excel(ws, 1, 1)
        Autofit(ws).autofit()
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # 将指针移回到文件的开头

        response = DownloadExcel.gen_response(output, "多环芳烃浓度2000-2030.xlsx")
        return response
